import io
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment

def html_to_excel_zip(html_content):
    soup = BeautifulSoup(html_content, "html.parser")

    wb = Workbook()
    ws = wb.active

    thin = Side(border_style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    occupied = {}

    row_idx = 1

    for tr in soup.find_all("tr"):
        col_idx = 1

        while (row_idx, col_idx) in occupied:
            col_idx += 1

        for td in tr.find_all(["td", "th"]):

            while (row_idx, col_idx) in occupied:
                col_idx += 1

            text = td.get_text(strip=True)
            colspan = int(td.get("colspan", 1))
            rowspan = int(td.get("rowspan", 1))

            cell = ws.cell(row=row_idx, column=col_idx, value=text)

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

            for r in range(row_idx, row_idx + rowspan):
                for c in range(col_idx, col_idx + colspan):
                    ws.cell(row=r, column=c).border = border

                    if (r, c) != (row_idx, col_idx):
                        occupied[(r, c)] = True

            if colspan > 1 or rowspan > 1:
                ws.merge_cells(
                    start_row=row_idx,
                    start_column=col_idx,
                    end_row=row_idx + rowspan - 1,
                    end_column=col_idx + colspan - 1
                )

            col_idx += colspan

        row_idx += 1

    # =========================
    # Save excel vào memory
    # =========================
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    return excel_buffer